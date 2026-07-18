#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from langgraph.types import Command
from openpyxl import load_workbook

from backend.app.agents.graph import (
    build_taxonomy_graph,
    create_initial_state,
    create_memory_checkpointer,
)
from backend.app.config import Settings
from backend.app.db import connect, init_db
from backend.app.repositories.file_repo import FileRepository
from backend.app.repositories.suggestion_repo import SuggestionRepository
from backend.app.repositories.task_repo import TaskRepository
from backend.app.services.action_service import ActionService
from backend.app.services.excel_service import UploadedFileMetadata
from scripts.generate_demo_excel import generate_excel


DEFAULT_GOLDEN_PATH = Path("eval/golden_issues.json")
DEFAULT_DEMO_PATH = Path("eval/golden_demo.xlsx")
DEFAULT_BASELINE_PATH = Path("eval/results/current_m1_m5_baseline.json")


@dataclass(frozen=True)
class GoldenIssue:
    issue_id: str
    issue_type: str
    affected_node_ids: tuple[int, ...]
    node_name: str
    expected_action: dict[str, Any]


@dataclass(frozen=True)
class GoldenSet:
    dataset_version: str
    source_excel: str
    description: str
    issues: tuple[GoldenIssue, ...]


def load_golden_issues(path: Path = DEFAULT_GOLDEN_PATH) -> GoldenSet:
    data = json.loads(path.read_text(encoding="utf-8"))
    issues = tuple(
        GoldenIssue(
            issue_id=str(item["issue_id"]),
            issue_type=str(item["issue_type"]),
            affected_node_ids=tuple(int(node_id) for node_id in item["affected_node_ids"]),
            node_name=str(item["node_name"]),
            expected_action=dict(item.get("expected_action") or {}),
        )
        for item in data["issues"]
    )
    return GoldenSet(
        dataset_version=str(data["dataset_version"]),
        source_excel=str(data["source_excel"]),
        description=str(data.get("description") or ""),
        issues=issues,
    )


def score_issues(golden: GoldenSet, predictions: list[dict[str, Any]]) -> dict[str, Any]:
    unmatched_golden = set(range(len(golden.issues)))
    matched_prediction_indexes: set[int] = set()
    matches: list[dict[str, Any]] = []

    for prediction_index, prediction in enumerate(predictions):
        for golden_index in sorted(unmatched_golden):
            issue = golden.issues[golden_index]
            if _prediction_matches_issue(prediction, issue):
                unmatched_golden.remove(golden_index)
                matched_prediction_indexes.add(prediction_index)
                matches.append(
                    {
                        "golden_issue_id": issue.issue_id,
                        "prediction_index": prediction_index,
                        "issue_type": issue.issue_type,
                    }
                )
                break

    overall = _metric_block(
        true_positive=len(matches),
        false_positive=len(predictions) - len(matched_prediction_indexes),
        false_negative=len(unmatched_golden),
    )
    by_type: dict[str, dict[str, Any]] = {}
    issue_types = sorted(
        {issue.issue_type for issue in golden.issues}
        | {str(item.get("issue_type")) for item in predictions}
    )
    for issue_type in issue_types:
        golden_indexes = {
            index for index, issue in enumerate(golden.issues) if issue.issue_type == issue_type
        }
        prediction_indexes = {
            index for index, item in enumerate(predictions) if item.get("issue_type") == issue_type
        }
        matched_predictions = {
            item["prediction_index"]
            for item in matches
            if item["issue_type"] == issue_type
        }
        matched_golden = {
            index
            for index, issue in enumerate(golden.issues)
            if issue.issue_type == issue_type
            and any(match["golden_issue_id"] == issue.issue_id for match in matches)
        }
        by_type[issue_type] = _metric_block(
            true_positive=len(matched_predictions),
            false_positive=len(prediction_indexes - matched_predictions),
            false_negative=len(golden_indexes - matched_golden),
        )

    return {
        "overall": overall,
        "by_type": by_type,
        "matches": matches,
        "missed_golden_issue_ids": [
            golden.issues[index].issue_id for index in sorted(unmatched_golden)
        ],
    }


def run_benchmark(
    *,
    golden_path: Path = DEFAULT_GOLDEN_PATH,
    demo_path: Path = DEFAULT_DEMO_PATH,
    output_path: Path = DEFAULT_BASELINE_PATH,
    use_live_llm: bool = False,
    approve_strategy: str = "low_risk",
) -> dict[str, Any]:
    golden = load_golden_issues(golden_path)
    if not demo_path.exists():
        demo_path.parent.mkdir(parents=True, exist_ok=True)
        generate_excel(str(demo_path))

    run_started_at = datetime.now().isoformat(timespec="seconds")
    with tempfile.TemporaryDirectory(prefix="taxonomy_benchmark_") as tmp:
        tmp_path = Path(tmp)
        settings = _benchmark_settings(tmp_path, use_live_llm=use_live_llm)
        settings.ensure_directories()
        init_db(settings)

        file_id = _create_file_record(settings, demo_path)
        task_repo = TaskRepository(settings)
        workflow_id = f"benchmark_{golden.dataset_version}_{int(time.time())}"
        thread_id = f"taxonomy_workflow:{workflow_id}"
        task_id = task_repo.create_workflow_task(
            file_id=file_id,
            workflow_id=workflow_id,
            thread_id=thread_id,
        )
        task_repo.record_event(
            workflow_id=workflow_id,
            thread_id=thread_id,
            task_id=task_id,
            node_name=None,
            event_type="workflow_started",
            status="running",
            progress=0,
            message="benchmark workflow started",
            payload={"file_id": file_id, "dataset_version": golden.dataset_version},
        )

        graph = build_taxonomy_graph(
            create_memory_checkpointer(),
            settings=settings,
            enable_suggestion_review=True,
        )
        state = create_initial_state(
            file_id=file_id,
            task_id=task_id,
            workflow_id=workflow_id,
        )
        config = {"configurable": {"thread_id": state.thread_id}}

        timings: dict[str, float] = {}
        node_outcomes: dict[str, dict[str, Any]] = {}
        final_state: dict[str, Any] = {}
        interrupted = False
        start = previous = time.perf_counter()
        for chunk in graph.stream(state, config=config, stream_mode="updates"):
            previous, interrupted = record_stream_chunk(
                chunk,
                timings,
                node_outcomes,
                final_state,
                previous,
            )
            if interrupted:
                break

        review_batch_id = final_state.get("review_batch_id")
        approved_ids: list[int] = []
        validation_summary = _empty_validation_summary()
        if interrupted and review_batch_id:
            suggestions = SuggestionRepository(settings).list_suggestions(
                review_batch_id=review_batch_id
            )
            approved_ids = _select_approved_suggestions(suggestions, approve_strategy)
            decision = {
                "decision": "approve" if approved_ids else "reject",
                "approved_suggestion_ids": approved_ids,
                "rejected_suggestion_ids": [
                    item.id for item in suggestions if item.id not in approved_ids
                ],
                "edits": [],
                "operator": "benchmark",
                "reject_reason": None if approved_ids else "benchmark rejected all suggestions",
            }
            for chunk in graph.stream(
                Command(resume=decision),
                config=config,
                stream_mode="updates",
            ):
                previous, _ = record_stream_chunk(
                    chunk,
                    timings,
                    node_outcomes,
                    final_state,
                    previous,
                )
            if approved_ids:
                validation_summary = _summarize_selected_validations(settings, approved_ids)

        elapsed_ms = round((time.perf_counter() - start) * 1000, 2)
        task = task_repo.get_task(task_id) or {}
        predictions = _list_diagnosis_issues(settings, final_state.get("base_version_id"))
        issue_metrics = score_issues(golden, predictions)
        suggestions = _list_suggestions(settings, final_state.get("base_version_id"))
        events = task_repo.list_events(workflow_id, after_id=0, limit=1000)

        preserved_report = preserve_artifact(
            Path(final_state["report_path"]) if final_state.get("report_path") else None,
            output_path.with_suffix(".report.md"),
        )
        result = {
            "benchmark_id": f"{golden.dataset_version}-{run_started_at}",
            "run_started_at": run_started_at,
            "dataset": {
                "version": golden.dataset_version,
                "source_excel": str(demo_path),
                "golden_issue_count": len(golden.issues),
                "row_count": _excel_row_count(demo_path),
            },
            "environment": {
                "use_live_llm": use_live_llm,
                "deepseek_configured": bool(settings.deepseek_api_key),
                "dashscope_configured": bool(settings.dashscope_api_key),
            },
            "workflow": {
                "workflow_id": workflow_id,
                "task_id": task_id,
                "status": task.get("status") or final_state.get("status"),
                "current_step": task.get("current_step") or final_state.get("current_step"),
                "measurement_scope": build_measurement_scope(use_live_llm=use_live_llm),
                "local_runner_elapsed_ms": elapsed_ms,
                "node_timings_ms": {key: round(value * 1000, 2) for key, value in timings.items()},
                "node_outcomes": node_outcomes,
                "failed_nodes": [
                    node_name
                    for node_name, outcome in node_outcomes.items()
                    if outcome.get("status") == "failed"
                ],
                "completed_steps": final_state.get("completed_steps", []),
                "interrupted_for_review": interrupted,
                "event_count": len(events),
            },
            "versions": {
                "base_version_id": final_state.get("base_version_id"),
                "current_version_id": final_state.get("current_version_id"),
                "new_version_id": final_state.get("new_version_id"),
                "version_no": final_state.get("version_no"),
            },
            "diagnosis": {
                "predicted_issue_count": len(predictions),
                "structure_issue_count": final_state.get("structure_issue_count", 0),
                "content_issue_count": final_state.get("content_issue_count", 0),
                "issue_metrics": issue_metrics,
            },
            "suggestions": {
                "suggestion_count": len(suggestions),
                "review_batch_id": review_batch_id,
                "approved_count": len(approved_ids),
                "action_types": _count_by_key(suggestions, "action_type"),
            },
            "actions": validation_summary,
            "artifacts": {
                "report_path": str(preserved_report) if preserved_report else None,
                "raw_report_path": final_state.get("report_path"),
                "output_path": str(output_path),
            },
        }
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(result, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return result


def _prediction_matches_issue(prediction: dict[str, Any], issue: GoldenIssue) -> bool:
    if prediction.get("issue_type") != issue.issue_type:
        return False
    prediction_node_id = prediction.get("node_id")
    if prediction_node_id is not None and int(prediction_node_id) in issue.affected_node_ids:
        return True
    text = " ".join(
        str(prediction.get(key) or "")
        for key in ("node_name", "description", "reason")
    )
    ids_in_text = {int(value) for value in re.findall(r"\d+", text)}
    if ids_in_text.intersection(issue.affected_node_ids):
        return True
    return issue.node_name and issue.node_name in text


def _metric_block(
    *,
    true_positive: int,
    false_positive: int,
    false_negative: int,
) -> dict[str, Any]:
    precision = true_positive / (true_positive + false_positive) if true_positive + false_positive else 0.0
    recall = true_positive / (true_positive + false_negative) if true_positive + false_negative else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
    return {
        "true_positive": true_positive,
        "false_positive": false_positive,
        "false_negative": false_negative,
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(f1, 4),
    }


def build_measurement_scope(*, use_live_llm: bool) -> dict[str, Any]:
    excludes = [
        "HTTP request/response latency",
        "SSE streaming latency",
        "frontend rendering time",
        "real human review waiting time",
    ]
    if not use_live_llm:
        excludes.extend(["LLM API latency", "embedding API latency", "remote Qdrant latency"])
    return {
        "timing_field": "local_runner_elapsed_ms",
        "includes": ["in-process LangGraph node execution"],
        "excludes": excludes,
        "valid_for_resume_metrics": True,
        "valid_for_user_wait_time": False,
    }


def _benchmark_settings(tmp_path: Path, *, use_live_llm: bool) -> Settings:
    return Settings(
        database_url=f"sqlite:///{tmp_path / 'app.db'}",
        upload_dir=tmp_path / "uploads",
        report_dir=tmp_path / "reports",
        export_dir=tmp_path / "exports",
        deepseek_api_key=os.getenv("DEEPSEEK_API_KEY", "") if use_live_llm else "",
        dashscope_api_key=os.getenv("DASHSCOPE_API_KEY", "") if use_live_llm else "",
    )


def _create_file_record(settings: Settings, source_path: Path) -> int:
    upload_path = settings.upload_dir / source_path.name
    shutil.copy2(source_path, upload_path)
    workbook = load_workbook(upload_path, read_only=True, data_only=True)
    sheet = workbook.active
    columns = [
        str(sheet.cell(row=1, column=index).value)
        for index in range(1, sheet.max_column + 1)
    ]
    metadata = UploadedFileMetadata(
        file_name=source_path.name,
        file_path=upload_path,
        file_size=upload_path.stat().st_size,
        sheet_name=sheet.title,
        row_count=max(sheet.max_row - 1, 0),
        column_count=sheet.max_column,
        columns=columns,
    )
    workbook.close()
    return FileRepository(settings).create_uploaded_file(metadata)


def record_stream_chunk(
    chunk: dict[str, Any],
    timings: dict[str, float],
    node_outcomes: dict[str, dict[str, Any]],
    final_state: dict[str, Any],
    previous: float,
    now: float | None = None,
) -> tuple[float, bool]:
    now = time.perf_counter() if now is None else now
    interrupted = False
    for node_name, update in chunk.items():
        if node_name == "__interrupt__":
            timings["wait_human_review_node"] = now - previous
            node_outcomes["wait_human_review_node"] = {
                "status": "interrupted",
                "current_step": "human_review",
            }
            interrupted = True
            continue
        timings[node_name] = now - previous
        if isinstance(update, dict):
            node_outcomes[node_name] = {
                key: update.get(key)
                for key in ("status", "current_step", "error_code", "error_message", "progress")
                if key in update
            }
            final_state.update(update)
    return now, interrupted


def _select_approved_suggestions(suggestions: list[Any], strategy: str) -> list[int]:
    if strategy == "none":
        return []
    if strategy == "all":
        return [int(item.id) for item in suggestions]
    return [
        int(item.id)
        for item in suggestions
        if item.risk_level == "low" and item.action_type == "mark_as_valid"
    ]


def _summarize_selected_validations(settings: Settings, suggestion_ids: list[int]) -> dict[str, Any]:
    suggestions = SuggestionRepository(settings).list_by_ids(suggestion_ids)
    validations = ActionService(settings).validate_suggestion_records(suggestions)
    return summarize_validation_results(validations)


def summarize_validation_results(validations: list[Any]) -> dict[str, Any]:
    passed = [item for item in validations if item.valid]
    failed = [item for item in validations if not item.valid]
    return {
        "validated_count": len(validations),
        "validation_passed_count": len(passed),
        "validation_failed_count": len(failed),
        "validation_pass_rate": round(len(passed) / len(validations), 4) if validations else 0.0,
        "failed_reasons": [item.reason for item in failed],
    }


def preserve_artifact(source: Path | None, target: Path) -> Path | None:
    if source is None or not source.exists():
        return None
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    return target


def _empty_validation_summary() -> dict[str, Any]:
    return {
        "validated_count": 0,
        "validation_passed_count": 0,
        "validation_failed_count": 0,
        "validation_pass_rate": 0.0,
        "failed_reasons": [],
    }


def _list_diagnosis_issues(settings: Settings, version_id: int | None) -> list[dict[str, Any]]:
    if version_id is None:
        return []
    with connect(settings) as connection:
        rows = connection.execute(
            """
            SELECT issue_type, node_id, node_name, description, reason, risk_level, confidence
            FROM diagnosis_issue
            WHERE version_id = ?
            ORDER BY id
            """,
            (version_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def _list_suggestions(settings: Settings, version_id: int | None) -> list[dict[str, Any]]:
    if version_id is None:
        return []
    with connect(settings) as connection:
        rows = connection.execute(
            """
            SELECT action_type, target_node_id, target_node_name, risk_level, confidence, status
            FROM adjustment_suggestion
            WHERE version_id = ?
            ORDER BY id
            """,
            (version_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def _count_by_key(items: list[dict[str, Any]], key: str) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in items:
        value = str(item.get(key) or "unknown")
        counts[value] = counts.get(value, 0) + 1
    return counts


def _excel_row_count(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return max(workbook.active.max_row - 1, 0)
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Run taxonomy agent benchmark on the pinned demo golden set.")
    parser.add_argument("--golden", type=Path, default=DEFAULT_GOLDEN_PATH)
    parser.add_argument("--demo", type=Path, default=DEFAULT_DEMO_PATH)
    parser.add_argument("--output", type=Path, default=DEFAULT_BASELINE_PATH)
    parser.add_argument("--use-live-llm", action="store_true", help="Use DeepSeek/DashScope keys from the environment.")
    parser.add_argument(
        "--approve-strategy",
        choices=("low_risk", "all", "none"),
        default="low_risk",
    )
    args = parser.parse_args()

    result = run_benchmark(
        golden_path=args.golden,
        demo_path=args.demo,
        output_path=args.output,
        use_live_llm=args.use_live_llm,
        approve_strategy=args.approve_strategy,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
