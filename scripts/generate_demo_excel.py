#!/usr/bin/env python3
"""
生成演示用的小型产品标准体系 Excel（~200 行）。
故意埋入系统可检测的各类问题，减少 demo 时的 token 消耗。

埋入的问题清单：
  结构诊断（规则检测）：
    1. missing_parent: 节点 150 的父节点 998 不存在
    2. deep_level:     节点 111/112 层级=8，超过阈值 7
    3. duplicate_name: "外套" 在 304/308/323 三处重复
  内容诊断（LLM 检测）：
    4. synonym_pollution: 103 苹果手机 syn_list 混入"华为手机"
    5. synonym_pollution: 104 华为手机 syn_list 混入"荣耀手机"
    6. synonym_pollution: 105 小米手机 syn_list 混入"iPhone"
    7. synonym_pollution: 132 电饭煲 syn_list 混入"电磁炉"
    8. synonym_pollution: 237 苹果(水果) syn_list 混入"iPhone"
    9. semantic_duplicate: 164 电子书阅读器 vs 165 电子书（不同分支同类义）
"""

import openpyxl
from openpyxl.styles import Font, Alignment

# ── 节点定义 ──────────────────────────────────────────────
# 格式: (category_id, category_name, parent_id, syn_list)
#   parent_id = None 表示根节点
#   syn_list  = None 或 "[]" 表示无同义词
#              否则为 '["词1", "词2"]' 格式的字符串

NODES: list[tuple[int, str, int | None, str | None]] = [
    # ═══════════════ 电子产品 (100) ═══════════════
    (100, "电子产品", None, None),

    # ── 手机 (101) ──
    (101, "手机", 100, None),
    (102, "智能手机", 101, None),
    (103, "苹果手机", 102, '["iPhone", "苹果", "华为手机"]'),       # synonym_pollution
    (104, "华为手机", 102, '["华为", "HUAWEI", "荣耀手机"]'),       # synonym_pollution
    (105, "小米手机", 102, '["小米", "红米手机", "iPhone"]'),       # synonym_pollution
    (106, "三星手机", 102, None),
    (107, "安卓手机", 102, None),
    (108, "国产安卓", 107, None),
    (109, "华为系列", 108, None),
    (110, "Mate系列", 109, None),
    (111, "Mate60", 110, None),       # deep_level (level=8)
    (112, "Mate60Pro", 110, None),    # deep_level (level=8)
    (113, "功能手机", 101, None),
    (114, "手机配件", 101, None),
    (115, "手机壳", 114, None),
    (116, "手机膜", 114, None),
    (117, "充电器", 114, None),
    (118, "数据线", 114, None),

    # ── 电脑 (119) ──
    (119, "电脑", 100, None),
    (120, "笔记本电脑", 119, None),
    (121, "商务笔记本", 120, None),
    (122, "游戏笔记本", 120, None),
    (123, "轻薄笔记本", 120, None),
    (124, "台式电脑", 119, None),
    (125, "电脑配件", 119, None),
    (126, "键盘", 125, None),
    (127, "鼠标", 125, None),
    (128, "显示器", 125, None),
    (129, "主机", 125, None),

    # ── 家用电器 (130) ──
    (130, "家用电器", 100, None),
    (131, "厨房电器", 130, None),
    (132, "电饭煲", 131, '["电饭锅", "高压锅", "电磁炉"]'),          # synonym_pollution
    (133, "微波炉", 131, None),
    (134, "榨汁机", 131, '["原汁机", "豆浆机", "搅拌机"]'),          # synonym_pollution
    (135, "电磁炉", 131, None),
    (136, "电水壶", 131, None),
    (137, "清洁电器", 130, None),
    (138, "吸尘器", 137, None),
    (139, "洗碗机", 137, None),
    (140, "扫地机器人", 137, None),
    (141, "个人护理", 130, None),
    (142, "电动牙刷", 141, None),
    (143, "电吹风", 141, None),
    (144, "剃须刀", 141, None),

    # ── 智能穿戴 (150) — 父节点 998 不存在 ──
    (150, "智能穿戴", 998, None),       # missing_parent
    (151, "智能手表", 150, None),
    (152, "智能手环", 150, None),
    (153, "智能眼镜", 150, None),

    # ── 音响设备 (154) ──
    (154, "音响设备", 100, None),
    (155, "蓝牙音箱", 154, None),
    (156, "家庭影院", 154, None),
    (157, "耳机", 154, None),
    (158, "有线耳机", 157, None),
    (159, "无线耳机", 157, None),

    # ── 摄影器材 (160) ──
    (160, "摄影器材", 100, None),
    (161, "数码相机", 160, None),
    (162, "摄像机", 160, None),
    (163, "运动相机", 160, None),

    # ── 游戏设备 (164) ──
    (164, "游戏设备", 100, None),
    (165, "游戏机", 164, None),
    (166, "游戏手柄", 164, None),
    (167, "游戏耳机", 164, None),

    # ── 电子阅读器 (168) ──
    (168, "电子阅读器", 100, None),
    (169, "电子书阅读器", 168, None),    # semantic_duplicate with 170
    (170, "电子书", 168, None),          # semantic_duplicate with 169

    # ═══════════════ 食品饮料 (200) ═══════════════
    (200, "食品饮料", None, None),

    # ── 休闲食品 (201) ──
    (201, "休闲食品", 200, None),
    (202, "膨化食品", 201, None),
    (203, "薯片", 202, None),
    (204, "爆米花", 202, None),
    (205, "锅巴", 202, None),
    (206, "坚果炒货", 201, None),
    (207, "瓜子", 206, None),
    (208, "花生", 206, None),
    (209, "核桃", 206, None),
    (210, "开心果", 206, None),
    (211, "糖果巧克力", 201, None),
    (212, "硬糖", 211, None),
    (213, "巧克力", 211, None),
    (214, "软糖", 211, None),
    (215, "饼干糕点", 201, None),
    (216, "饼干", 215, None),
    (217, "蛋糕", 215, None),
    (218, "面包", 215, None),

    # ── 饮料 (219) ──
    (219, "饮料", 200, None),
    (220, "碳酸饮料", 219, None),
    (221, "果汁饮料", 219, None),
    (222, "茶饮料", 219, None),
    (223, "矿泉水", 219, None),
    (224, "咖啡", 219, None),
    (225, "奶制品", 219, None),
    (226, "牛奶", 225, None),
    (227, "酸奶", 225, None),
    (228, "奶粉", 225, None),

    # ── 粮油调味 (229) ──
    (229, "粮油调味", 200, None),
    (230, "大米", 229, None),
    (231, "食用油", 229, None),
    (232, "调味品", 229, None),
    (233, "酱油", 232, None),
    (234, "醋", 232, None),
    (235, "盐", 232, None),

    # ── 水果 (236) ──
    (236, "水果", 200, None),
    (237, "苹果", 236, '["Apple", "苹果", "iPhone"]'),             # synonym_pollution
    (238, "香蕉", 236, None),
    (239, "橙子", 236, None),
    (240, "葡萄", 236, None),

    # ── 海鲜水产 (241) ──
    (241, "海鲜水产", 200, None),
    (242, "鱼类", 241, None),
    (243, "虾类", 241, None),
    (244, "贝类", 241, None),

    # ── 肉禽蛋 (245) ──
    (245, "肉禽蛋", 200, None),
    (246, "猪肉", 245, None),
    (247, "牛肉", 245, None),
    (248, "鸡肉", 245, None),
    (249, "鸡蛋", 245, None),

    # ── 茶叶 (250) ──
    (250, "茶叶", 200, None),
    (251, "绿茶", 250, None),
    (252, "红茶", 250, None),
    (253, "乌龙茶", 250, None),
    (254, "花茶", 250, None),

    # ═══════════════ 服装鞋帽 (300) ═══════════════
    (300, "服装鞋帽", None, None),

    # ── 男装 (301) ──
    (301, "男装", 300, None),
    (302, "衬衫", 301, None),
    (303, "裤子", 301, None),
    (304, "外套", 301, None),            # duplicate_name (1/3)

    # ── 女装 (305) ──
    (305, "女装", 300, None),
    (306, "连衣裙", 305, None),
    (307, "半身裙", 305, None),
    (308, "外套", 305, None),            # duplicate_name (2/3)
    (309, "打底衫", 305, None),

    # ── 鞋类 (310) ──
    (310, "鞋类", 300, None),
    (311, "运动鞋", 310, None),
    (312, "皮鞋", 310, None),
    (313, "凉鞋", 310, None),
    (314, "靴子", 310, None),

    # ── 配饰 (315) ──
    (315, "配饰", 300, None),
    (316, "帽子", 315, None),
    (317, "围巾", 315, None),
    (318, "腰带", 315, None),
    (319, "手套", 315, None),

    # ── 童装 (320) ──
    (320, "童装", 300, None),
    (321, "童装上衣", 320, None),
    (322, "童装裤子", 320, None),
    (323, "外套", 320, None),            # duplicate_name (3/3)

    # ── 运动服装 (324) ──
    (324, "运动服装", 300, None),
    (325, "运动上衣", 324, None),
    (326, "运动裤", 324, None),
    (327, "运动套装", 324, None),

    # ── 内衣 (328) ──
    (328, "内衣", 300, None),
    (329, "文胸", 328, None),
    (330, "内裤", 328, None),
    (331, "袜子", 328, None),

    # ═══════════════ 家居用品 (400) ═══════════════
    (400, "家居用品", None, None),

    # ── 家具 (401) ──
    (401, "家具", 400, None),
    (402, "客厅家具", 401, None),
    (403, "沙发", 402, None),
    (404, "茶几", 402, None),
    (405, "卧室家具", 401, None),
    (406, "床", 405, None),
    (407, "衣柜", 405, None),
    (408, "书房家具", 401, None),
    (409, "书桌", 408, None),
    (410, "书架", 408, None),

    # ── 厨房用品 (411) ──
    (411, "厨房用品", 400, None),
    (412, "餐具", 411, None),
    (413, "锅具", 411, None),
    (414, "刀具", 411, None),

    # ── 卫浴用品 (415) ──
    (415, "卫浴用品", 400, None),
    (416, "毛巾", 415, None),
    (417, "浴巾", 415, None),
    (418, "牙刷", 415, None),

    # ── 床品 (419) ──
    (419, "床品", 400, None),
    (420, "床单", 419, None),
    (421, "被套", 419, None),
    (422, "枕头", 419, None),

    # ── 灯具 (423) ──
    (423, "灯具", 400, None),
    (424, "台灯", 423, None),
    (425, "吊灯", 423, None),
    (426, "落地灯", 423, None),

    # ── 装饰品 (427) ──
    (427, "装饰品", 400, None),
    (428, "花瓶", 427, None),
    (429, "挂画", 427, None),
    (430, "摆件", 427, None),

    # ── 清洁用品 (431) ──
    (431, "清洁用品", 400, None),
    (432, "洗衣液", 431, None),
    (433, "洗洁精", 431, None),
    (434, "垃圾袋", 431, None),
]


def build_node_dict(nodes: list[tuple]) -> dict[int, tuple]:
    """构建 id -> (name, parent_id, syn_list) 映射"""
    return {node[0]: (node[1], node[2], node[3]) for node in nodes}


def get_ancestor_ids(node_id: int, node_dict: dict) -> list[int]:
    """获取节点的所有祖先 ID（从根到直接父节点）。
    如果父节点不存在于 node_dict 中（missing_parent 场景），
    仍然将其加入 ancestors 列表。"""
    ancestors = []
    current = node_id
    while True:
        if current not in node_dict:
            break
        _, parent_id, _ = node_dict[current]
        if parent_id is None:
            break
        ancestors.insert(0, parent_id)
        current = parent_id
        if len(ancestors) > 20:
            break
    return ancestors


def get_ancestor_names(ancestor_ids: list[int], node_dict: dict) -> list[str]:
    """获取祖先名称列表。不存在的祖先用占位符。"""
    return [
        node_dict[aid][0] if aid in node_dict else f"(未知节点:{aid})"
        for aid in ancestor_ids
    ]


def format_category_group_id(ancestor_ids: list[int]) -> str | None:
    """格式化 category_group_id: '1,2,3' 或 None（根节点）"""
    if not ancestor_ids:
        return None
    return ",".join(str(aid) for aid in ancestor_ids)


def format_category_pids(ancestor_ids: list[int]) -> str:
    """格式化 category_pids: '[-1],[2],[3]' 或 '[-1],'（根节点）"""
    if not ancestor_ids:
        return "[-1],"
    parts = ["[-1]"] + [f"[{aid}]" for aid in ancestor_ids]
    return ",".join(parts)


def format_category_group_name(ancestor_names: list[str]) -> str | None:
    """格式化 category_group_name: '电子产品,手机,智能手机' 或 None（根节点）"""
    if not ancestor_names:
        return None
    return ",".join(ancestor_names)


def format_syn_list(syn_list: str | None) -> str:
    """格式化 syn_list"""
    if syn_list is None:
        return "[]"
    return syn_list


def generate_excel(output_path: str) -> None:
    """生成 Excel 文件"""
    node_dict = build_node_dict(NODES)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头
    headers = [
        "category_id",
        "category_name",
        "category_group_id",
        "category_pids",
        "category_group_name",
        "syn_list",
    ]
    header_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, (cat_id, cat_name, parent_id, syn_list) in enumerate(NODES, 2):
        ancestor_ids = get_ancestor_ids(cat_id, node_dict)
        ancestor_names = get_ancestor_names(ancestor_ids, node_dict)

        ws.cell(row=row_idx, column=1, value=cat_id)
        ws.cell(row=row_idx, column=2, value=cat_name)
        ws.cell(row=row_idx, column=3, value=format_category_group_id(ancestor_ids))
        ws.cell(row=row_idx, column=4, value=format_category_pids(ancestor_ids))
        ws.cell(row=row_idx, column=5, value=format_category_group_name(ancestor_names))
        ws.cell(row=row_idx, column=6, value=format_syn_list(syn_list))

    # 调整列宽
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 30

    wb.save(output_path)
    print(f"已生成: {output_path}")
    print(f"总行数: {len(NODES)} 行数据 + 1 行表头")

    # 打印埋入的问题清单
    print("\n══════════ 埋入的问题清单 ══════════")
    print("\n【结构诊断 — 规则检测】")
    print("  1. missing_parent:  节点 150「智能穿戴」父节点 998 不存在")
    print("  2. deep_level:      节点 111「Mate60」层级=8 (链路: 100→101→102→107→108→109→110→111)")
    print("  3. deep_level:      节点 112「Mate60Pro」层级=8 (同上链路)")
    print("  4. duplicate_name:  「外套」在 304(男装)/308(女装)/323(童装) 三处重复")

    print("\n【内容诊断 — LLM 检测】")
    print("  5. synonym_pollution: 103「苹果手机」syn_list 含「华为手机」(竞品品牌)")
    print("  6. synonym_pollution: 104「华为手机」syn_list 含「荣耀手机」(子品牌不是同义词)")
    print("  7. synonym_pollution: 105「小米手机」syn_list 含「iPhone」(竞品品牌)")
    print("  8. synonym_pollution: 132「电饭煲」syn_list 含「电磁炉」(不同品类)")
    print("  9. synonym_pollution: 134「榨汁机」syn_list 含「搅拌机」(不同品类)")
    print(" 10. synonym_pollution: 237「苹果」(水果) syn_list 含「iPhone」(电子产品)")
    print(" 11. semantic_duplicate: 169「电子书阅读器」vs 170「电子书」(同一父节点下语义重复)")

    print(f"\n总节点数: {len(NODES)}")
    print(f"有同义词的节点数: {sum(1 for n in NODES if n[3] and n[3] != '[]')}")
    print(f"根节点数: {sum(1 for n in NODES if n[2] is None)}")


if __name__ == "__main__":
    output = "/Users/flflfl/Documents/code/SystemMaintenanceAgent/data/sample/产品标准体系_demo.xlsx"
    generate_excel(output)
