from openpyxl import load_workbook

from backend.app.config import Settings
from backend.app.db import init_db
from backend.app.repositories.diagnosis_repo import DiagnosisRepository
from backend.app.repositories.suggestion_repo import SuggestionRepository
from backend.app.repositories.taxonomy_repo import TaxonomyRepository
from backend.app.repositories.version_repo import VersionRepository
from backend.app.schemas.issue import DiagnosisIssueRecord
from backend.app.schemas.suggestion import AdjustmentSuggestion
from backend.app.schemas.taxonomy import TaxonomyNodeRecord
from backend.app.services.report_service import ReportService
from backend.app.tools.export_tools import export_excel


def _settings(tmp_path):
    return Settings(
        database_url=f"sqlite:///{tmp_path / 'app.db'}",
        upload_dir=tmp_path / "uploads",
        report_dir=tmp_path / "reports",
        export_dir=tmp_path / "exports",
        deepseek_api_key="",
        dashscope_api_key="",
    )


def _seed_version(settings: Settings) -> int:
    init_db(settings)
    version_id = VersionRepository(settings).create_version(
        file_id=1,
        version_no="v1.0",
        description="test",
        quality_score=95.0,
    )
    TaxonomyRepository(settings).bulk_insert_nodes(
        version_id=version_id,
        nodes=[
            TaxonomyNodeRecord(
                category_id=1,
                category_name="根",
                parent_id=None,
                level=1,
                path_ids="1",
                path_names="根",
                category_group_id=None,
                category_pids=None,
                category_group_name=None,
                syn_list=None,
                is_leaf=0,
            ),
            TaxonomyNodeRecord(
                category_id=2,
                category_name="苹果",
                parent_id=1,
                level=2,
                path_ids="1,2",
                path_names="根 > 苹果",
                category_group_id="1",
                category_pids="1",
                category_group_name="根",
                syn_list="红富士",
                is_leaf=1,
            ),
        ],
    )
    issue_id = DiagnosisRepository(settings).create_issue(
        version_id=version_id,
        issue=DiagnosisIssueRecord(
            issue_type="synonym_pollution",
            node_id=2,
            node_name="苹果",
            description="同义词污染",
            reason="混入非水果词",
            risk_level="medium",
            confidence=0.8,
        ),
    )
    SuggestionRepository(settings).create_suggestion(
        review_batch_id="batch-report",
        suggestion=AdjustmentSuggestion(
            issue_id=issue_id,
            version_id=version_id,
            action_type="clean_synonym",
            target_node_id=2,
            target_node_name="苹果",
            action_payload={"synonyms_to_remove": ["AirPods"]},
            reason="测试原因",
            suggestion="删除污染词",
            risk_level="medium",
            confidence=0.9,
            need_confirm=True,
            status="approved",
        ),
    )
    return version_id


def test_export_excel_writes_standard_columns(tmp_path):
    settings = _settings(tmp_path)
    version_id = _seed_version(settings)

    export_path = export_excel(version_id, settings)

    assert export_path.name == "v1.0_taxonomy.xlsx"
    workbook = load_workbook(export_path, read_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == (
        "category_id",
        "category_name",
        "category_group_id",
        "category_pids",
        "category_group_name",
        "syn_list",
    )
    assert rows[2] == (2, "苹果", "1", "1", "根", "红富士")


def test_m4_report_contains_suggestions_execution_and_version_sections(tmp_path):
    settings = _settings(tmp_path)
    version_id = _seed_version(settings)

    result = ReportService(settings).generate_diagnosis_report(version_id)
    report_text = result.report_path.read_text(encoding="utf-8")

    assert "## 6. 智能维护建议" in report_text
    assert "删除污染词" in report_text
    assert "## 7. 版本变更记录" in report_text
    assert "## 8. 质量评分" in report_text
    assert "95.0/100" in report_text
