import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from fastapi import UploadFile
from openpyxl import load_workbook

from backend.app.config import Settings


REQUIRED_COLUMNS = {
    "category_id",
    "category_name",
    "category_group_id",
    "category_pids",
    "category_group_name",
    "syn_list",
}


class ExcelValidationError(ValueError):
    def __init__(self, message: str, error_code: str) -> None:
        super().__init__(message)
        self.error_code = error_code


@dataclass(frozen=True)
class UploadedFileMetadata:
    file_name: str
    file_path: Path
    file_size: int
    sheet_name: str
    row_count: int
    column_count: int
    columns: list[str]


@dataclass(frozen=True)
class ParseExcelResult:
    file_id: int
    file_name: str
    file_path: Path
    sheet_name: str
    row_count: int
    column_count: int
    columns: list[str]


class ExcelService:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    async def save_and_inspect(self, upload_file: UploadFile) -> UploadedFileMetadata:
        original_name = upload_file.filename or "uploaded.xlsx"
        suffix = Path(original_name).suffix.lower()
        if suffix not in self.settings.allowed_upload_suffixes:
            raise ExcelValidationError("Only .xlsx files are supported.", "INVALID_FILE_TYPE")

        self.settings.upload_dir.mkdir(parents=True, exist_ok=True)
        saved_path = self.settings.upload_dir / self._saved_file_name(original_name)

        with saved_path.open("wb") as output:
            shutil.copyfileobj(upload_file.file, output)

        file_size = saved_path.stat().st_size
        if file_size == 0:
            saved_path.unlink(missing_ok=True)
            raise ExcelValidationError("Excel file is empty.", "EMPTY_FILE")
        if file_size > self.settings.max_upload_size_bytes:
            saved_path.unlink(missing_ok=True)
            raise ExcelValidationError("Excel file is too large.", "FILE_TOO_LARGE")

        try:
            workbook = load_workbook(saved_path, read_only=True, data_only=True)
        except Exception as exc:
            saved_path.unlink(missing_ok=True)
            raise ExcelValidationError("Excel file could not be read.", "INVALID_EXCEL") from exc

        try:
            sheet = workbook.worksheets[0]
            columns, row_count = _inspect_worksheet(sheet)
            missing_columns = sorted(REQUIRED_COLUMNS.difference(columns))
            if missing_columns:
                joined = ", ".join(missing_columns)
                raise ExcelValidationError(
                    f"Excel missing required columns: {joined}",
                    "INVALID_COLUMNS",
                )
            sheet_name = sheet.title
        except ExcelValidationError:
            workbook.close()
            saved_path.unlink(missing_ok=True)
            raise
        finally:
            workbook.close()

        return UploadedFileMetadata(
            file_name=original_name,
            file_path=saved_path,
            file_size=file_size,
            sheet_name=sheet_name,
            row_count=row_count,
            column_count=len(columns),
            columns=columns,
        )

    def parse_uploaded_file(self, file_id: int) -> ParseExcelResult:
        from backend.app.repositories.file_repo import FileRepository

        file_record = FileRepository(self.settings).get_file(file_id)
        if file_record is None:
            raise ExcelValidationError("Uploaded file was not found.", "FILE_NOT_FOUND")

        file_path = Path(file_record["file_path"])
        if not file_path.exists():
            raise ExcelValidationError("Uploaded file path does not exist.", "FILE_NOT_FOUND")

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ExcelValidationError("Excel file could not be read.", "INVALID_EXCEL") from exc

        try:
            sheet = workbook.worksheets[0]
            columns, row_count = _inspect_worksheet(sheet)
            missing_columns = sorted(REQUIRED_COLUMNS.difference(columns))
            if missing_columns:
                joined = ", ".join(missing_columns)
                raise ExcelValidationError(
                    f"Excel missing required columns: {joined}",
                    "INVALID_COLUMNS",
                )
            sheet_name = sheet.title
        finally:
            workbook.close()

        return ParseExcelResult(
            file_id=file_id,
            file_name=file_record["file_name"],
            file_path=file_path,
            sheet_name=sheet_name,
            row_count=row_count,
            column_count=len(columns),
            columns=columns,
        )

    def _saved_file_name(self, original_name: str) -> str:
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(original_name).name)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
        return f"{timestamp}_{safe_name}"


def _inspect_worksheet(sheet) -> tuple[list[str], int]:
    """Read headers and count populated data rows without relying on dimensions.

    Some valid XLSX writers omit worksheet dimension metadata, leaving
    ``ReadOnlyWorksheet.max_row`` as ``None``. Iterating also handles those files.
    """
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration as exc:
        raise ExcelValidationError("Excel worksheet is empty.", "EMPTY_FILE") from exc
    columns = ["" if value is None else str(value).strip() for value in header]
    row_count = sum(1 for row in rows if any(value is not None for value in row))
    return columns, row_count
