from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from backend.app.config import Settings
from backend.app.repositories.file_repo import FileRepository
from backend.app.repositories.taxonomy_repo import TaxonomyRepository
from backend.app.repositories.version_repo import VersionRepository
from backend.app.schemas.taxonomy import (
    BuildTreeResult,
    OverviewResult,
    TaxonomyNodeRecord,
)
from backend.app.services.excel_service import REQUIRED_COLUMNS, ExcelValidationError


class TaxonomyService:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    def build_tree(self, file_id: int) -> BuildTreeResult:
        nodes = self.parse_tree_nodes(file_id)
        return self._summarize(file_id, nodes)

    def get_overview(self, version_id: int) -> OverviewResult:
        version = VersionRepository(self.settings).get_version(version_id)
        if version is None:
            raise ValueError(f"Taxonomy version {version_id} was not found.")
        counts = TaxonomyRepository(self.settings).get_overview_counts(version_id)
        return OverviewResult(
            version_id=version_id,
            file_id=int(version["file_id"]),
            node_count=counts["node_count"],
            root_count=counts["root_count"],
            max_depth=counts["max_depth"],
            max_children_count=counts["max_children_count"],
            leaf_count=counts["leaf_count"],
            non_leaf_count=counts["non_leaf_count"],
            missing_parent_count=counts["missing_parent_count"],
            duplicate_name_count=counts["duplicate_name_count"],
            synonym_non_empty_count=counts["synonym_non_empty_count"],
        )

    def parse_tree_nodes(self, file_id: int) -> list[TaxonomyNodeRecord]:
        file_record = FileRepository(self.settings).get_file(file_id)
        if file_record is None:
            raise ExcelValidationError("Uploaded file was not found.", "FILE_NOT_FOUND")
        rows = self._read_rows(Path(file_record["file_path"]))
        raw_nodes = [self._row_to_node(row) for row in rows]
        category_ids = [node.category_id for node in raw_nodes]
        duplicates = [item for item, count in Counter(category_ids).items() if count > 1]
        if duplicates:
            raise ValueError(f"Duplicate category_id values: {duplicates[:5]}")
        parent_ids = {node.parent_id for node in raw_nodes if node.parent_id is not None}
        leaf_ids = set(category_ids).difference(parent_ids)
        return [
            node.model_copy(update={"is_leaf": 1 if node.category_id in leaf_ids else 0})
            for node in raw_nodes
        ]

    def _read_rows(self, file_path: Path) -> list[dict[str, Any]]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        iterator = sheet.iter_rows(values_only=True)
        columns = ["" if value is None else str(value).strip() for value in next(iterator)]
        missing_columns = sorted(REQUIRED_COLUMNS.difference(columns))
        if missing_columns:
            joined = ", ".join(missing_columns)
            raise ExcelValidationError(f"Excel missing required columns: {joined}", "INVALID_COLUMNS")
        rows = []
        for values in iterator:
            row = dict(zip(columns, values, strict=False))
            if all(value is None for value in row.values()):
                continue
            rows.append(row)
        return rows

    def _row_to_node(self, row: dict[str, Any]) -> TaxonomyNodeRecord:
        category_id = _parse_required_int(row.get("category_id"), "category_id")
        category_name = _clean_text(row.get("category_name"))
        if not category_name:
            raise ValueError(f"category_name is empty for category_id {category_id}.")
        ancestor_ids = _split_ints(row.get("category_group_id"))
        parent_id = ancestor_ids[-1] if ancestor_ids else None
        group_names = _split_names(row.get("category_group_name"))
        path_ids = [*ancestor_ids, category_id]
        path_names = [*group_names, category_name]
        return TaxonomyNodeRecord(
            category_id=category_id,
            category_name=category_name,
            parent_id=parent_id,
            level=len(ancestor_ids) + 1,
            path_ids=",".join(str(item) for item in path_ids),
            path_names=" > ".join(path_names),
            category_group_id=_clean_optional_text(row.get("category_group_id")),
            category_pids=_clean_optional_text(row.get("category_pids")),
            category_group_name=_clean_optional_text(row.get("category_group_name")),
            syn_list=_clean_optional_text(row.get("syn_list")),
        )

    def _summarize(
        self,
        file_id: int,
        nodes: list[TaxonomyNodeRecord],
    ) -> BuildTreeResult:
        category_ids = {node.category_id for node in nodes}
        child_counts: dict[int, int] = defaultdict(int)
        for node in nodes:
            if node.parent_id is not None:
                child_counts[node.parent_id] += 1
        name_counts = Counter(node.category_name for node in nodes)
        return BuildTreeResult(
            file_id=file_id,
            node_count=len(nodes),
            root_count=sum(1 for node in nodes if node.parent_id is None),
            max_depth=max((node.level for node in nodes), default=0),
            max_children_count=max(child_counts.values(), default=0),
            leaf_count=sum(1 for node in nodes if node.is_leaf == 1),
            non_leaf_count=sum(1 for node in nodes if node.is_leaf == 0),
            missing_parent_count=sum(
                1
                for node in nodes
                if node.parent_id is not None and node.parent_id not in category_ids
            ),
            duplicate_name_count=sum(1 for count in name_counts.values() if count > 1),
            synonym_non_empty_count=sum(
                1 for node in nodes if node.syn_list not in {None, "", "[]"}
            ),
        )


def _parse_required_int(value: Any, field_name: str) -> int:
    if value is None or value == "":
        raise ValueError(f"{field_name} is required.")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be an integer.") from exc


def _split_ints(value: Any) -> list[int]:
    text = _clean_text(value)
    if not text:
        return []
    return [int(part.strip()) for part in text.split(",") if part.strip()]


def _split_names(value: Any) -> list[str]:
    text = _clean_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clean_optional_text(value: Any) -> str | None:
    text = _clean_text(value)
    return text if text else None
